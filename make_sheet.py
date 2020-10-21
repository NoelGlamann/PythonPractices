#!usr/bin/python3
#Noel Glamann
#December 12, 2019

'''
Create and populate an slxs file called 'users.xlsx'.
'''

import openpyxl
import datetime as dt
import time
first = ""
last = ""
uname = ""
workbook = openpyxl.Workbook()
sheet = workbook.create_sheet('User Data')

#set headings row
sheet['A1'].value = "Time Stamp"
sheet['B1'].value = "First Name"
sheet['C1'].value = "Last Name"
sheet['D1'].value = "Username"

for i in range(6):
    new_row = sheet.max_row + 1
    sheet.cell(row = sheet.max_row + 1, column = 1).value = dt.datetime.now()
    first = input("Enter First Name: ")
    last = input("Enter Last Name: ")
    uname = input("Select User Name: ")
    sheet.cell(row = new_row, column = 2).value = first
    sheet.cell(row = new_row, column = 3).value = last
    sheet.cell(row = new_row, column = 4).value = uname
    
workbook.save('users.xlsx')